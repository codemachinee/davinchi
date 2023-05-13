# библиотека работы с гугл таблицами
# import gspread
import openai
# библиотека проверки даты
from datetime import datetime
from yoomoney import Client, Quickpay
from openpyxl import load_workbook  # библиотека работы с exel таблицами
from paswords import *
from keyboards import *

saved_messages_davinci = []


async def Davinci(bot, message, text, answer_message, proverka):
    global statistic
    try:
        openai.api_key = Davinci_key
        answer_davinci = open('davinci.txt.txt', 'r', encoding='utf-8')
        saved_messages_davinci.insert(0, f'Вы: {text}\n')
        prompt_davinci = (str(answer_davinci.read()) + ''.join(reversed(saved_messages_davinci)))
        if proverka == 'YES':
            response = openai.Completion.create(
                model="text-davinci-003",
                prompt=prompt_davinci,
                temperature=0.0,
                max_tokens=1000,
                top_p=1,
                frequency_penalty=0.0,
                presence_penalty=0.0,
            )
            await bot.edit_message_text(f'{response["choices"][0]["text"]}', message.chat.id, answer_message.message_id)
            saved_messages_davinci.insert(0, f'{str(response["choices"][0]["text"])}\n')
            if len(saved_messages_davinci) >= 8:
                del saved_messages_davinci[3:]
        else:
            await bot.edit_message_text(f'Для дальнейшего использования бота нужно пополнить баланс', message.chat.id,
                                        answer_message.message_id)
            await buttons(bot, message).oplata_buttons(url=await platezhy(bot, message).url_generation())
    except Exception:
        await bot.send_message(message.chat.id, "Простите но мне нужен перекур..")
        await bot.send_message(admin_id, "Простите но мне нужен перекур..")
        del saved_messages_davinci[1:]


class statistic:
    def __init__(self):
        self.wb = load_workbook('chek_list.xlsx')
        self.ws = self.wb['посещения']

    def obnulenie(self):
        seconds_now = (datetime.now().time().hour * 3600 + datetime.now().time().minute * 60 + datetime.now().time().second)
        for row in self.ws['B2':f'C{self.ws.max_row}']:
            if row[1].value != 0:
                row[1].value = int(row[1].value) - 1
                if 0 <= seconds_now <= 21600:
                    row[0].value = 0
            else:
                row[0].value = 0
        self.wb.save('chek_list.xlsx')

    async def proverka(self, message):
        for row in self.ws['A2':f'C{self.ws.max_row}']:
            if row[0].value == message.chat.id:
                if row[2].value != 0:
                    return 'YES'
                else:
                    if row[1].value <= 2:
                        row[1].value += 1
                        self.wb.save('chek_list.xlsx')
                        return 'YES'
                    else:
                        return "NO"
        new_row = self.ws.max_row + 1
        self.ws[f'A{new_row}'].value = message.chat.id
        self.ws[f'B{new_row}'].value = 1
        self.ws[f'C{new_row}'].value = 0
        self.wb.save('chek_list.xlsx')
        return 'YES'

    async def status(self, message):
        for row in self.ws['A2':f'C{self.ws.max_row}']:
            if row[0].value == message.chat.id:
                if row[2].value == 0:
                    return f'Подписка отсутствует. Использовано {row[1].value} из 3 бесплатных запросов'
                elif 0 < row[2].value <= 4:
                    return f'Оформлен безлимит на сутки. Осталось более {row[2].value * 6} часов пользования.'
                elif row[2].value > 4:
                    return f'Оформлен недельный безлимит. Осталось более {row[2].value * 6} часов пользования.'
        return f'Подписка отсутствует. Использовано 0 из 3 бесплатных запросов'


class platezhy:
    def __init__(self, bot, message):
        self.bot = bot
        self.message = message
        try:
            self.marker_mess = str(self.message.chat.id)
        except AttributeError:
            self.marker_mess = str(self.message.message.chat.id)

    async def url_generation(self):
        try:
            quickpay = Quickpay(
                receiver="4100116460956966",
                quickpay_form="shop",
                targets="payment",
                paymentType="SB",
                sum=10,
                label=self.marker_mess
            )
            return quickpay.base_url
        except AttributeError:
            quickpay = Quickpay(
                receiver="4100116460956966",
                quickpay_form="shop",
                targets="payment",
                paymentType="SB",
                sum=10,
                label=self.marker_mess
            )
            return quickpay.base_url

    async def chec_control(self):
        wb = load_workbook('chek_list.xlsx')
        ws = wb['посещения']
        token = token_umany
        client = Client(token)
        try:
            history = client.operation_history(label=self.marker_mess)
            # for operation in history.operations:
            #     print()
            #     print("Operation:", operation.operation_id)
            #     print("\tStatus     -->", operation.status)
            #     print("\tDatetime   -->", operation.datetime)
            #     print("\tTitle      -->", operation.title)
            #     print("\tPattern id -->", operation.pattern_id)
            #     print("\tDirection  -->", operation.direction)
            #     print("\tAmount     -->", operation.amount)
            # print(history.operations[0].amount)
        except AttributeError:
            history = client.operation_history(label=self.marker_mess)
        try:
            if (datetime.now().day == history.operations[0].datetime.day) or (datetime.now().day ==
                                                                              history.operations[0].datetime.day + 1):
                if history.operations[0].amount == 9.7:
                    for row in ws['A2':f'C{ws.max_row}']:
                        if row[0].value == self.message.message.chat.id:
                            if row[2].value == 0:
                                row[1].value = 0
                                row[2].value = 4
                                await self.bot.send_message(self.message.message.chat.id, f'Оплата прошла, спасибо! '
                                                                                          f'У вас 24 часа пользования '
                                                                                          f'ботом.')
                                await self.bot.send_message(admin_id, f'🚨!!!ВНИМАНИЕ!!!🚨\n'
                                                                      f'Поступила оплата от:\n'
                                                                      f'id чата: {self.message.message.chat.id}\n'
                                                                      f'Имя: {self.message.from_user.first_name}\n'
                                                                      f'Фамилия: {self.message.from_user.last_name}\n'
                                                                      f'Ссылка: @{self.message.from_user.username}\n')
                                wb.save('chek_list.xlsx')
                                break
                            else:
                                await self.bot.send_message(self.message.message.chat.id, f'Платеж был проверен ранее, '
                                                                                          f'у вас 24 часа пользования '
                                                                                          f'ботом.')
                                break
                elif history.operations[0].amount > 9.7:
                    for row in ws['A2':f'C{ws.max_row}']:
                        if row[0].value == self.message.message.chat.id:
                            if row[2].value == 0:
                                row[1].value = 0
                                row[2].value = 28
                                await self.bot.send_message(self.message.message.chat.id, f'Оплата прошла, спасибо! '
                                                                                          f'У вас 7 дней пользования '
                                                                                          f'ботом.')
                                await self.bot.send_message(admin_id, f'🚨!!!ВНИМАНИЕ!!!🚨\n'
                                                                      f'Поступила оплата от:\n'
                                                                      f'id чата: {self.message.message.chat.id}\n'
                                                                      f'Имя: {self.message.from_user.first_name}\n'
                                                                      f'Фамилия: {self.message.from_user.last_name}\n'
                                                                      f'Ссылка: @{self.message.from_user.username}\n')
                                wb.save('chek_list.xlsx')
                                break
                            elif 1 <= row[2].value <= 4:
                                row[1].value = 0
                                row[2].value += 28
                                await self.bot.send_message(self.message.message.chat.id, f'Оплата прошла, спасибо! '
                                                                                          f'У вас добавлено 7 дней '
                                                                                          f'пользования ботом.')
                                await self.bot.send_message(admin_id, f'🚨!!!ВНИМАНИЕ!!!🚨\n'
                                                                      f'Поступила доплата от:\n'
                                                                      f'id чата: {self.message.message.chat.id}\n'
                                                                      f'Имя: {self.message.from_user.first_name}\n'
                                                                      f'Фамилия: {self.message.from_user.last_name}\n'
                                                                      f'Ссылка: @{self.message.from_user.username}\n')
                                wb.save('chek_list.xlsx')
                                break
                            elif row[2].value >= 24:
                                await self.bot.send_message(self.message.message.chat.id, f'Платеж был проверен ранее, '
                                                                                          f'у вас 24 часа пользования '
                                                                                          f'ботом.')
            else:
                await self.bot.send_message(self.message.message.chat.id, f'Платеж не найден. '
                                                                          f'Если Вы оплатили товар, напишите в '
                                                                          f'поддержку @hloapps')
                await buttons(self.bot, self.message).oplata_buttons(url=await platezhy(self.bot, self.message).url_generation())
        except IndexError:
            await self.bot.send_message(self.message.message.chat.id, 'Платеж не найден. Если Вы оплатили товар, '
                                                                      'напишите в поддержку @hloapps')
            await buttons(self.bot, self.message).oplata_buttons(url=await platezhy(self.bot, self.message).url_generation())



