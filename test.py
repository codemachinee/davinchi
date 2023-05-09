from openpyxl import load_workbook


def obnulenie():
    wb = load_workbook('chek_list.xlsx')
    ws = wb['посещения']

    # for row in ws['B2':f'C{ws.max_row}']:
    #     if row[1].value != 0:
    #         row[1].value = int(row[1].value) - 1
    #         row[0].value = 0
    #     else:
    #         row[0].value = 0
    # wb.save('chek_list.xlsx')
    print(ws.max_row)
    new_row = ws.max_row + 1
    ws[f'A{new_row}'].value = 1
    ws[f'B{new_row}'].value = 2
    ws[f'C{new_row}'].value = 3
    wb.save('chek_list.xlsx')
    print(ws.max_row)


obnulenie()